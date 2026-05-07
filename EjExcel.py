import os
#Archivo de excel archivo.xlsx
import openpyxl
#Workbook o libro de excel (la diagonal para usar en la dirección es / ):
wb=openpyxl.load_workbook("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/PROING26.xlsx") 
#WorkSheet u Hoja activa de Excel
#ws = wb.active 


def ppal(): 
    celdas_usadas=0
    ws=wb["PROING"] #Nombre de la hoja
    CantMAXfilas=ws.max_row #conteo de filas
    CantMAXcol=ws.max_column #Conteo de columnas
    print("Filas: ",CantMAXfilas," Columnas: ",CantMAXcol)
    cell = ws['A2'] #Evitamos encabezado
    print("Datos")
    for i in range(2,CantMAXfilas+1): #Quitamos el encabezado por eso comienza en 2
      #Este for navega en las celdas para saber si hay datos 
      
      if cell.value is not None:
         valor=ws.cell(row=i, column=1).value
         print("Valor en celda ",i,": ",valor)
         celdas_usadas+=1
      else:
         print("La celda está vacía.")
    print("Celdas usadas= ",celdas_usadas)

    #Ya sabemos cuales celdas son usadas
    #Para leer datos del excel
    print("Matricula|   Nombres        |       Apellidos")
    for i in range(2,CantMAXfilas+1):
       matricula=ws.cell(row=i, column=1).value
       nombres=ws.cell(row=i, column=2).value
       apellidos=ws.cell(row=i, column=3).value
       print(matricula,"    ",nombres,"          ",apellidos)
    input("Oprime una tecla para continuar")
    os.system('cls' if os.name == 'nt' else 'clear')
     
   #Agregar valores a una celda
    print(" Agregar calificaciones: ")
    print("  ")
    for i in range(2,CantMAXfilas+1):
       matricula=ws.cell(row=i, column=1).value
       nombres=ws.cell(row=i, column=2).value
       apellidos=ws.cell(row=i, column=3).value
       print("Ingresa las calificaciones para ",nombres," ",apellidos)
       primer=int(input("Ingresa la calificación del primer parcial: "))
       segundo=int(input("Ingresa la calificación del segundo parcial: "))
       tercer=int(input("Ingresa la calificación del tercer parcial: "))
       ws.cell(row=i, column=4, value=primer) #columna 4, se asigna el valor de primer
       ws.cell(row=i, column=5, value=segundo) #columna 5,se asigna el valor de segundo
       ws.cell(row=i, column=6, value=tercer) #columna 6,se asigna el valor de tercer
       ws.cell(row=i, column=7, value=(primer*0.3 + segundo*0.3 +tercer*0.4))
       
    wb.save("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/PROING26.xlsx")#se guardan los cambios en el archivo
    input("Oprime una tecla para continuar")
    os.system('cls' if os.name == 'nt' else 'clear')

      #Muestra calificaciones
    print("Matricula|   Nombres        |       Apellidos              |  1P  |  2P  | 3P | PROMEDIO |")
    for i in range(2,CantMAXfilas+1):
      matricula=ws.cell(row=i, column=1).value
      nombres=ws.cell(row=i, column=2).value
      apellidos=ws.cell(row=i, column=3).value
      uno=ws.cell(row=i, column=4).value
      dos=ws.cell(row=i, column=5).value
      tres=ws.cell(row=i, column=6).value
      final=ws.cell(row=i, column=7).value
      print(matricula,"    ",nombres,"          ",apellidos,"      ",uno,"      ",dos,"     ",tres, "    ",final)
       
   

if __name__=="__main__":
  ppal()  