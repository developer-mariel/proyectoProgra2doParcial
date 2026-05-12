import os
import time
import openpyxl
from colorama import init, Fore, Back, Style
init()

# variables globales
x = 0
wb = ""
CantMAXfilas = 0
CantMAXcol = 0
fila = 0
matricula = 0
Nombre = " "
Apellido = " "

# Función para abrir el archivo de excel
def abrir_excel():
    global x, wb
    try:
        wb = openpyxl.load_workbook("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/EjemsExc/PROING26.xlsx")
        x = 1
        print("Archivo abierto con éxito")
    except:
        print("Algo se puso raro...\U0001F914")
        x = 0
    return x


def conteoFC(ws):
    global CantMAXfilas, CantMAXcol, fila
    CantMAXfilas = ws.max_row
    CantMAXcol = ws.max_column
    print(f"Cantidad de filas: {CantMAXfilas} Cantidad de columnas {CantMAXcol}")
    print("La hoja MATERIA tiene ", CantMAXfilas, "filas y ", CantMAXcol, " columnas")
    print(" ")
    input("Oprime una tecla para continuar")
    os.system('cls' if os.name == 'nt' else 'clear')


def buscarIntegrante(ws):
    global matricula, Nombre, Apellido, fila
    print(" ")
    encontradoN = "No"
    matricula = int(input("Ingrese la matricula: "))

    for i in range(2, CantMAXfilas + 1):
        Mat = ws.cell(row=i, column=1).value
        Matint = int(Mat)

        if Matint == matricula:
            print(" \U0001F9D0 ")
            fila = i
            print("El alumno se encontró, ya existe en el archivo")
            encontradoN = "Si"
            Nombre = ws.cell(row=i, column=2).value
            Apellido = ws.cell(row=i, column=3).value
            print(f"Nombre: {Nombre}, apellido(s): {Apellido}")
            continuar()

    return encontradoN


def eliminarIntegrante(ws):
    print(" ")
    ws.delete_rows(fila)
    print(f"Se eliminó el registro {matricula} {Nombre} {Apellido}")


def continuar():
    input("Oprime una tecla para continuar")
    os.system('cls' if os.name == 'nt' else 'clear')


def ppal():
    Edo = abrir_excel()

    if Edo == 1:
        opc = 0
        agregados = 0

        while opc != 6:
            print(Fore.YELLOW + " ADMINISTRACIÓN DE EQUIPOS EN EXCEL ", "\n"
                  "1.EQUIPO", "\n"
                  "2.BUSQUEDA POR ALUMNO", "\n"
                  "3.CALIFICACIONES Y PROMEDIO DEL EQUIPO", "\n"
                  "4.PERMISO PARA EL NO ORDINARIO POR EQUIPO", "\n"
                  "5.EQUIPAZO", "\n"
                  "6.SALIR")

            opc = int(input("Ingresa una opción: "))
            print(" " + Style.RESET_ALL)

            match opc:

                case 1:
                    print(Back.BLUE + "1.CREAR EQUIPO" + Style.RESET_ALL)
                    ws = wb["PROING"]
                    conteoFC(ws)

                    print("1. Agregar un integrante al equipo")
                    print("2. Eliminar un integrante del equipo")
                    eleccion = int(input("Ingrese la opción: "))

                    if eleccion == 1:
                        if agregados >= 2:
                            print("Solo puedes agregar máximo 2 integrantes")
                            continuar()
                        else:
                            encontrado = buscarIntegrante(ws)

                            if encontrado == "No":
                                print(" \U0001F913 ")
                                print("Datos del nuevo integrante")
                                Nombre = input("Ingrese el o los nombres: ")
                                Apellido = input("Ingrese los apellidos: ")

                                ws.cell(row=CantMAXfilas + 1, column=1, value=matricula)
                                ws.cell(row=CantMAXfilas + 1, column=2, value=Nombre)
                                ws.cell(row=CantMAXfilas + 1, column=3, value=Apellido)

                                wb.save("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/EjemsExc/PROING26.xlsx")

                                agregados += 1
                                print("Alumno agregado correctamente")
                                continuar()

                    elif eleccion == 2:
                        alumnos = CantMAXfilas - 1

                        if alumnos <= 1:
                            print("No puedes eliminar, debe quedar al menos un integrante")
                            continuar()
                        else:
                            encontrado = buscarIntegrante(ws)

                            if encontrado == "Si":
                                eliminarIntegrante(ws)
                                wb.save("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/EjemsExc/PROING26.xlsx")
                                continuar()
                            else:
                                print("No se encontró el integrante")
                                continuar()

                case 2:
                    print(Back.BLUE + "2.BUSQUEDA POR ALUMNO" + Style.RESET_ALL)
                    ws = wb["PROING"]
                    conteoFC(ws)

                    encontrado = buscarIntegrante(ws)

                    if encontrado == "No":
                        print("No se encontró el alumno")
                        resp = input("¿Deseas agregarlo? (S/N): ")

                        if resp == "S":
                            if agregados <= 2:
                                print("Datos del nuevo integrante")
                                Nombre = input("Ingrese el o los nombres: ")
                                Apellido = input("Ingrese los apellidos: ")

                                ws.cell(row=CantMAXfilas + 1, column=1, value=matricula)
                                ws.cell(row=CantMAXfilas + 1, column=2, value=Nombre)
                                ws.cell(row=CantMAXfilas + 1, column=3, value=Apellido)

                                wb.save("C:/Users/ianterrazasrugelio/Documents/Trabajos de la universidad/Segundo semestre/PROGRAM/Python/3ER PARCIAL/EjemsExc/PROING26.xlsx")

                                agregados += 1
                                print("Alumno agregado correctamente")
                                continuar()
                            else:
                                print("Solo puedes agregar máximo 2 integrantes")
                                continuar()
                        else:
                            print("Regresando al menú...")
                            continuar()

                case 3:
                    print(Back.BLUE + "3.CALIFICACIONES Y PROMEDIO DEL EQUIPO" + Style.RESET_ALL)

                case 4:
                    print(Back.BLUE + "4.PERMISO PARA EL NO ORDINARIO POR EQUIPO" + Style.RESET_ALL)

                case 5:
                    print(Back.BLUE + "5.EQUIPAZO" + Style.RESET_ALL)

                case 6:
                    print(Back.BLUE + "6.SALIR" + Style.RESET_ALL)

    else:
        print("El archivo no está creado o está abierto \U0001F6A7")


if __name__ == "__main__":
    ppal()